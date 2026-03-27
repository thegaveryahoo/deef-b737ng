#!/usr/bin/env python3
import sys, os, zipfile, base64, json
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Installeer eerst: pip install openpyxl pillow")
    sys.exit(1)

def extract_images(xlsx_path):
    images = []
    with zipfile.ZipFile(xlsx_path, 'r') as z:
        media = sorted([f for f in z.namelist() if f.startswith('xl/media/')])
        for mf in media:
            data = z.read(mf)
            ext = mf.split('.')[-1].lower()
            mime = 'image/png' if ext == 'png' else 'image/jpeg'
            b64 = base64.b64encode(data).decode()
            images.append(f'data:{mime};base64,{b64}')
    return images

def extract_text(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    cards = []
    for sheet in wb.worksheets:
        texts = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and str(cell.value).strip():
                    texts.append(str(cell.value).strip())
        if texts:
            cards.append({'sheet': sheet.title, 'texts': texts})
    return cards

def build_app(images, text_data, out_path):
    flashcards = []
    img_i = 0
    for page in text_data:
        txts = page['texts']
        mid = len(txts) // 2
        fronts = txts[:mid] if mid > 0 else txts
        backs = txts[mid:] if mid > 0 else ['']
        num = max(1, min(5, len(backs)))
        for i in range(num):
            img = images[img_i] if img_i < len(images) else None
            img_i += 1
            flashcards.append({
                'id': len(flashcards),
                'sheet': page['sheet'],
                'front_img': img,
                'front_text': fronts[i] if i < len(fronts) and not img else '',
                'back': backs[i] if i < len(backs) else ''
            })
    while img_i < len(images):
        flashcards.append({'id': len(flashcards), 'sheet': 'Extra',
            'front_img': images[img_i], 'front_text': '', 'back': ''})
        img_i += 1

    cj = json.dumps(flashcards, ensure_ascii=False)
    
    html = open(os.path.join(os.path.dirname(__file__), 'template.html')).read() if os.path.exists('template.html') else None
    
    page = """<!DOCTYPE html>
<html lang="nl"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>B737NG-800 Flashcards</title>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:system-ui,sans-serif;background:#f0f2f5;padding:12px;min-height:100vh}
.app{max-width:680px;margin:0 auto}
.hdr{background:#1a1a2e;color:#fff;border-radius:12px;padding:14px 18px;margin-bottom:12px;display:flex;justify-content:space-between;align-items:center}
.hdr h1{font-size:15px;font-weight:500;margin:0}
.hdr .sub{font-size:11px;opacity:.6;margin-top:2px}
.stats{display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:12px}
.stat{background:#fff;border-radius:8px;padding:10px;text-align:center;border:0.5px solid #e8e8e8}
.sn{font-size:20px;font-weight:500}.sl{font-size:11px;color:#888;margin-top:1px}
.good .sn{color:#1a7a4a}.mwah .sn{color:#a07000}.bad .sn{color:#c02020}
.prog-wrap{margin-bottom:12px}
.prog-labels{display:flex;justify-content:space-between;font-size:12px;color:#888;margin-bottom:4px}
.prog-bar{background:#e8e8e8;border-radius:4px;height:6px}
.prog-fill{background:#1a1a2e;height:6px;border-radius:4px;transition:width .4s}
.modes{display:flex;gap:8px;margin-bottom:12px}
.mode-btn{flex:1;padding:8px;border-radius:6px;border:0.5px solid #ddd;background:#fff;font-size:12px;cursor:pointer;color:#555}
.mode-btn.active{background:#1a1a2e;color:#fff;border-color:#1a1a2e}
.card-wrap{background:#fff;border-radius:12px;border:0.5px solid #e8e8e8;overflow:hidden;margin-bottom:12px}
.card-front{background:#eef2ff;min-height:200px;display:flex;flex-direction:column;align-items:center;justify-content:center;padding:20px;border-bottom:0.5px solid #d8dff0}
.card-front img{max-width:100%;max-height:220px;object-fit:contain;border-radius:6px}
.front-lbl{font-size:10px;letter-spacing:1px;color:#8898cc;text-transform:uppercase;margin-bottom:10px}
.front-txt{font-size:17px;font-weight:500;color:#1a1a2e;text-align:center;line-height:1.5}
.card-back{padding:18px;min-height:90px}
.reveal-wrap{display:flex;justify-content:center;padding:20px}
.reveal-btn{background:#1a1a2e;color:#fff;border:none;border-radius:8px;padding:12px 32px;font-size:14px;cursor:pointer}
.back-txt{font-size:14px;color:#222;line-height:1.8;white-space:pre-wrap}
.score-row{display:flex;gap:8px;margin-top:14px}
.sbtn{flex:1;padding:11px 6px;border-radius:8px;border:none;font-size:13px;font-weight:500;cursor:pointer}
.sbtn:active{transform:scale(.97)}
.s-good{background:#e6f4ec;color:#1a7a4a;border:1px solid #b8ddc8}
.s-mwah{background:#fef8e6;color:#8a6000;border:1px solid #f0d890}
.s-bad{background:#faeaea;color:#c02020;border:1px solid #f0b8b8}
.nav{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px}
.nav-btn{background:#fff;border:0.5px solid #ddd;border-radius:6px;padding:7px 16px;font-size:13px;cursor:pointer}
.badge{display:inline-block;font-size:11px;padding:2px 8px;border-radius:10px}
.b-good{background:#e6f4ec;color:#1a7a4a}
.b-mwah{background:#fef8e6;color:#8a6000}
.b-bad{background:#faeaea;color:#c02020}
.b-new{background:#eeeeff;color:#5544cc}
.done{text-align:center;padding:40px 20px;background:#fff;border-radius:12px}
.done h2{font-size:22px;margin-bottom:8px}
.rbtn{background:#1a1a2e;color:#fff;border:none;border-radius:8px;padding:12px 24px;font-size:14px;cursor:pointer;margin:8px 4px 0}
.rbtn.sec{background:#555}
</style></head><body>
<div class="app">
<div class="hdr">
  <div><h1>B737NG-800 Flashcards</h1><div class="sub">Transavia gezagvoerder opleiding</div></div>
  <span id="ctr" style="font-size:13px;color:#aac"></span>
</div>
<div class="stats">
  <div class="stat"><div class="sn" id="cnt-tot">""" + str(len(flashcards)) + """</div><div class="sl">Totaal</div></div>
  <div class="stat good"><div class="sn" id="cnt-good">0</div><div class="sl">Good</div></div>
  <div class="stat mwah"><div class="sn" id="cnt-mwah">0</div><div class="sl">Mwah</div></div>
  <div class="stat bad"><div class="sn" id="cnt-bad">0</div><div class="sl">Bad</div></div>
</div>
<div class="prog-wrap">
  <div class="prog-labels"><span id="prog-lbl">0 van """ + str(len(flashcards)) + """ beheerst</span><span id="prog-pct">0%</span></div>
  <div class="prog-bar"><div class="prog-fill" id="prog-fill" style="width:0%"></div></div>
</div>
<div class="modes">
  <button class="mode-btn active" id="m-due" onclick="setMode('due')">Nu te herhalen</button>
  <button class="mode-btn" id="m-all" onclick="setMode('all')">Alle kaarten</button>
  <button class="mode-btn" id="m-weak" onclick="setMode('weak')">Alleen zwak</button>
</div>
<div id="main">
  <div class="nav">
    <button class="nav-btn" onclick="prev()">&#8592; Vorige</button>
    <span id="badge"></span>
    <button class="nav-btn" onclick="next()">Volgende &#8594;</button>
  </div>
  <div class="card-wrap">
    <div class="card-front"><div class="front-lbl">Voorzijde</div><div id="front"></div></div>
    <div class="card-back" id="cback">
      <div class="reveal-wrap" id="rw"><button class="reveal-btn" onclick="reveal()">Toon achterzijde</button></div>
      <div id="bd" style="display:none">
        <div class="back-txt" id="btxt"></div>
        <div class="score-row">
          <button class="sbtn s-good" onclick="rate('good')">Good<br><small style="opacity:.7">+7d</small></button>
          <button class="sbtn s-mwah" onclick="rate('mwah')">Mwah<br><small style="opacity:.7">+3d</small></button>
          <button class="sbtn s-bad" onclick="rate('bad')">Bad<br><small style="opacity:.7">+1d</small></button>
        </div>
      </div>
    </div>
  </div>
</div>
<div id="done" class="done" style="display:none">
  <div style="font-size:48px">&#10003;</div>
  <h2>Sessie klaar!</h2>
  <p style="color:#666;font-size:14px;margin-top:8px">Alle kaarten voor nu doorlopen.</p>
  <div class="stats" style="margin-top:20px">
    <div class="stat"><div class="sn" id="d-tot">0</div><div class="sl">Totaal</div></div>
    <div class="stat good"><div class="sn" id="d-good">0</div><div class="sl">Good</div></div>
    <div class="stat mwah"><div class="sn" id="d-mwah">0</div><div class="sl">Mwah</div></div>
    <div class="stat bad"><div class="sn" id="d-bad">0</div><div class="sl">Bad</div></div>
  </div>
  <p id="nxt-txt" style="color:#888;font-size:13px;margin-top:16px"></p>
  <button class="rbtn" onclick="resetAll()">Alles opnieuw</button>
  <button class="rbtn sec" onclick="setMode('all')">Toch alle kaarten</button>
</div>
</div>
<script>
const CARDS=""" + cj + """;
const KEY='b737fc_v2';
const IV={good:7,mwah:3,bad:1};
const EV={good:1.5,mwah:1.0,bad:0.5};
let st=load(),queue=[],qi=0,mode='due';
function load(){try{const r=localStorage.getItem(KEY);if(r)return JSON.parse(r);}catch(e){}return{cards:{}};}
function save(){try{localStorage.setItem(KEY,JSON.stringify(st));}catch(e){}}
function gc(id){if(!st.cards[id])st.cards[id]={score:'new',interval:1,nextDate:null,reps:0};return st.cards[id];}
function isDue(id){const c=gc(id);if(!c.nextDate)return true;return new Date(c.nextDate)<=new Date();}
function buildQ(){
  if(mode==='due') queue=CARDS.filter(c=>isDue(c.id)).map(c=>c.id);
  else if(mode==='weak') queue=CARDS.filter(c=>['bad','mwah'].includes(gc(c.id).score)).map(c=>c.id);
  else queue=CARDS.map(c=>c.id);
  const o={bad:0,mwah:1,new:2,good:3};
  queue.sort((a,b)=>(o[gc(a).score]||2)-(o[gc(b).score]||2));
  qi=0;
}
function render(){
  updateStats();
  if(!queue.length||qi>=queue.length){showDone();return;}
  document.getElementById('main').style.display='block';
  document.getElementById('done').style.display='none';
  const id=queue[qi],card=CARDS[id],cs=gc(id);
  document.getElementById('ctr').textContent=(qi+1)+' / '+queue.length;
  const f=document.getElementById('front');
  if(card.front_img) f.innerHTML='<img src="'+card.front_img+'" alt="Flashcard">';
  else f.innerHTML='<div class="front-txt">'+esc(card.front_text||card.sheet)+'</div>';
  document.getElementById('btxt').textContent=card.back;
  document.getElementById('rw').style.display='flex';
  document.getElementById('bd').style.display='none';
  const sc=cs.score||'new';
  const bl={good:'b-good',mwah:'b-mwah',bad:'b-bad',new:'b-new'};
  const sl={good:'Good',mwah:'Mwah',bad:'Bad',new:'Nieuw'};
  document.getElementById('badge').innerHTML='<span class="badge '+(bl[sc]||'b-new')+'">'+(sl[sc]||'Nieuw')+'</span>';
  updateProg();
}
function reveal(){document.getElementById('rw').style.display='none';document.getElementById('bd').style.display='block';}
function rate(score){
  const id=queue[qi],cs=gc(id);
  cs.score=score; cs.reps=(cs.reps||0)+1;
  let iv=cs.interval||1;
  iv=Math.max(1,Math.round(iv*EV[score]));
  if(score==='bad')iv=1;
  if(score==='mwah')iv=Math.max(3,Math.min(iv,5));
  if(score==='good')iv=Math.max(7,iv);
  cs.interval=iv;
  const nd=new Date(); nd.setDate(nd.getDate()+iv);
  cs.nextDate=nd.toISOString();
  save(); qi++; render();
}
function next(){if(qi<queue.length-1){qi++;render();}}
function prev(){if(qi>0){qi--;render();}}
function setMode(m){
  mode=m;
  ['due','all','weak'].forEach(n=>document.getElementById('m-'+n).className='mode-btn'+(n===m?' active':''));
  buildQ();
  document.getElementById('done').style.display='none';
  document.getElementById('main').style.display='block';
  render();
}
function updateStats(){
  const sc=Object.values(st.cards).map(c=>c.score);
  document.getElementById('cnt-good').textContent=sc.filter(s=>s==='good').length;
  document.getElementById('cnt-mwah').textContent=sc.filter(s=>s==='mwah').length;
  document.getElementById('cnt-bad').textContent=sc.filter(s=>s==='bad').length;
}
function updateProg(){
  const good=Object.values(st.cards).filter(c=>c.score==='good').length;
  const pct=Math.round(100*good/CARDS.length);
  document.getElementById('prog-fill').style.width=pct+'%';
  document.getElementById('prog-pct').textContent=pct+'%';
  document.getElementById('prog-lbl').textContent=good+' van '+CARDS.length+' beheerst';
}
function showDone(){
  document.getElementById('main').style.display='none';
  document.getElementById('done').style.display='block';
  const sc=Object.values(st.cards).map(c=>c.score);
  document.getElementById('d-tot').textContent=CARDS.length;
  document.getElementById('d-good').textContent=sc.filter(s=>s==='good').length;
  document.getElementById('d-mwah').textContent=sc.filter(s=>s==='mwah').length;
  document.getElementById('d-bad').textContent=sc.filter(s=>s==='bad').length;
  const nds=Object.values(st.cards).filter(c=>c.nextDate).map(c=>new Date(c.nextDate)).sort((a,b)=>a-b);
  if(nds.length){const d=Math.ceil((nds[0]-new Date())/86400000);
    document.getElementById('nxt-txt').textContent=d<=0?'Er zijn al kaarten klaar!':'Volgende kaarten over '+d+' dag(en).';}
}
function resetAll(){if(confirm('Alle voortgang wissen en opnieuw beginnen?')){st={cards:{}};save();setMode('all');}}
function esc(s){return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}
buildQ(); render();
<\/script></body></html>"""
    
    with open(out_path, 'w', encoding='utf-8') as f:
        f.write(page)
    print(f"\nKlaar! Geopend als: {out_path}")
    print(f"Afbeeldingen: {len(images)}, Flashcards: {len(flashcards)}")

if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else 'CMD_flashcards_v2_0.xlsx'
    if not os.path.exists(xlsx):
        print(f"Bestand niet gevonden: {xlsx}"); sys.exit(1)
    print(f"Lezen: {xlsx}...")
    imgs = extract_images(xlsx)
    print(f"  {len(imgs)} afbeeldingen gevonden")
    txts = extract_text(xlsx)
    print(f"  {len(txts)} werkbladen gevonden")
    build_app(imgs, txts, 'flashcards_app.html')
